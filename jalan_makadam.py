import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def jalan_makadam_flow():
    st.subheader("Jalan Makadam")

    # Menambahkan 2 kolom untuk gambar di bagian atas
    col1, col2 = st.columns(2)

    with col1:
        st.image("assets/jalan makadam 1.jpg", caption="1", use_column_width=True)

    with col2:
        st.image("assets/jalan makadam 2.jpg", caption="2", use_column_width=True)

    # Tombol untuk memulai input
    if st.button("Mulai Input"):
        st.session_state.jalan_makadam = {}
        st.session_state.show_galian_input = True
        st.session_state.show_urugan_input = False
        st.session_state.show_estimasi_input = False

    # Bagian Galian
    if st.session_state.get('show_galian_input', False):
        st.write("### Pekerjaan Galian")
        jenis_galian = st.selectbox("Jenis Galian", ["Galian Batu", "Galian Tanah", "Galian Lumpur", "Galian Pasir", "Galian Cadas"])
        #metode = st.radio("Metode Pekerjaan", ["Manual", "Mekanis", "Semi Mekanis"])

# Tentukan metode berdasarkan jenis galian yang dipilih

        if jenis_galian in ["Galian Batu", "Galian Tanah", "Galian Lumpur"]:
            metode = st.radio("Metode Pekerjaan", ["Manual", "Mekanis", "Semi Mekanis"], key="metode_full_options")
        elif jenis_galian in ["Galian Pasir", "Galian Cadas"]:
            metode = st.radio("Metode Pekerjaan", ["Manual", "Semi Mekanis"], key="metode_limited_options")

        panjang = st.number_input("Panjang Galian (m)", format="%.2f", min_value=0.0)
        lebar = st.number_input("Lebar Galian (m)", format="%.2f", min_value=0.0)
        kedalaman = st.number_input("Kedalaman Makadam (m)", format="%.2f", min_value=0.0)


        if st.button("Konfirmasi Galian", key="konfirmasi_galian"):
            if panjang == 0 or lebar == 0 or kedalaman == 0:
                st.error("Mohon masukkan panjang, lebar, dan kedalaman yang valid sebelum melanjutkan.")
            else:
                st.session_state.jalan_makadam.update({
                    'pekerjaan': "Galian",
                    'jenis_galian': jenis_galian,
                    'metode': metode,
                    'panjang_galian': panjang,
                    'lebar_galian': lebar,
                    'kedalaman_galian': kedalaman,
                })
                st.success("Input Galian disimpan! Lanjutkan ke Urugan.")
                st.session_state.show_urugan_input = True
                # st.session_state.show_galian_input = False

    # Bagian Urugan
    if st.session_state.get('show_urugan_input', False):
        st.write("### Pekerjaan Pemasangan Jalan Makadam")
        panjang_urugan = st.number_input("Panjang Urugan (m)", format="%.2f", min_value=0.0)
        lebar_urugan = st.number_input("Lebar Urugan (m)", format="%.2f", min_value=0.0)
        kedalaman_urugan = st.number_input("Kedalaman Urugan (m)", format="%.2f", min_value=0.0)
        pemadatan_urugan = st.selectbox("Jenis Pemadatan", ["Stamper", "Bulldozer"])

        if st.button("Konfirmasi Urugan", key="konfirmasi_urugan"):
            if panjang_urugan == 0 or lebar_urugan == 0:
                st.error("Mohon masukkan panjang dan lebar yang valid sebelum melanjutkan.")
            else:
                st.session_state.jalan_makadam.update({
                    'panjang_urugan': panjang_urugan,
                    'lebar_urugan': lebar_urugan,
                    'kedalaman_urugan': kedalaman_urugan,
                    'pemadatan_urugan': pemadatan_urugan,
                })
                st.success("Input Urugan disimpan! Tekan tombol Submit untuk menyimpan semua data.")
                st.session_state.show_estimasi_input = True
                # st.session_state.show_urugan_input = False
 
            volume_galian = panjang * lebar * kedalaman
            luas_urugan = panjang_urugan * lebar_urugan * kedalaman_urugan
            volume_pemadatan = panjang_urugan * lebar_urugan * kedalaman_urugan
        
            # Inisialisasi data
            data = {}

            # Definisi data untuk tabel berdasarkan jenis_galian dan metode
            if jenis_galian == 'Galian Batu':
                if metode == 'Manual':
                    data_galian = {
                        'Uraian': ['Galian', 'Pekerja', 'Mandor'],
                        'Koefisien': [np.nan, 3.378, 0.3378],
                        'Volume': [np.nan] + [volume_galian] * 2,  # Menggunakan np.nan untuk baris pertama
                        'Satuan': [np.nan, 'OH', 'OH'],
                        'Harga Satuan': [np.nan, 81500, 98000],
                        'Jumlah': [np.nan] * 3
                    }
                    st.session_state.data_galian = data_galian
                elif metode == 'Semi Mekanis':
                    data_galian = {
                        'Uraian': ['Galian', 'Pertamina dex', 'Jack Hammer', 'Pekerja', 'Mandor'],
                        'Koefisien': [np.nan, 2.5, 0.25, 1, 0.1],
                        'Volume': [np.nan] + [volume_galian] * 4,  # Menggunakan np.nan untuk baris pertama
                        'Satuan': [np.nan, 'liter', 'sewa-harian', 'OH', 'OH'],
                        'Harga Satuan': [np.nan, 13300, 44700, 81500, 98000],
                        'Jumlah': [np.nan] * 5
                    }
                    st.session_state.data_galian = data_galian
                elif metode == 'Mekanis':
                    data_galian = {
                        'Uraian': ['Galian', 'Compressor', 'Jack Hammer', 'Wheel Loader', 'Excavator', 'Dump Truck', 'Pekerja (Buruh Tidak Terampil)', 'Mandor'],
                        'Koefisien': [np.nan, 0.0738, 0.0738, 0.0738, 0.0738, 0.3329, 0.0843, 0.0105],
                        'Volume': [np.nan] + [volume_galian] * 7,  # Menggunakan np.nan untuk baris pertama
                        'Satuan': [np.nan, 'jam', 'jam','jam','jam','jam', 'OH', 'OH'],
                        'Harga Satuan': [np.nan, 212427.45, 38465.57, 538209.99, 819402.32, 339612.94, 81500, 98000],
                        'Jumlah': [np.nan] * 8
                    }
                    st.session_state.data_galian = data_galian
            elif jenis_galian == 'Galian Tanah':
                if metode == 'Manual':
                    data_galian = {
                        'Uraian': ['Galian', 'Pekerja', 'Mandor'],
                        'Koefisien': [np.nan, 0.563, 0.0675],
                        'Volume': [np.nan] + [volume_galian] * 2,  # Menggunakan np.nan untuk baris pertama
                        'Satuan': [np.nan, 'OH', 'OH'],
                        'Harga Satuan': [np.nan, 81500, 98000],
                        'Jumlah': [np.nan] * 3
                    }
                    st.session_state.data_galian = data_galian
                elif metode == 'Semi Mekanis':
                    data_galian = {
                        'Uraian': ['Galian', 'Pertamina dex', 'Jack Hammer', 'Pekerja', 'Mandor'],
                        'Koefisien': [np.nan, 0.5, 0.05, 0.22, 0.022],
                        'Volume': [np.nan] + [volume_galian] * 4,  # Menggunakan np.nan untuk baris pertama
                        'Satuan': [np.nan, 'liter', 'sewa-harian', 'OH', 'OH'],
                        'Harga Satuan': [np.nan, 13300, 44700, 81500, 98000],
                        'Jumlah': [np.nan] * 5
                    }
                    st.session_state.data_galian = data_galian
                if metode == 'Mekanis':
                    data_galian = {
                        'Uraian': ['Galian', 'Excavator','Pekerja', 'Mandor'],
                        'Koefisien': [np.nan, 0.0414, 0.83, 0.083],
                        'Volume': [np.nan] + [volume_galian] * 3,
                        'Satuan': [np.nan, 'jam', 'OH', 'OH'],
                        'Harga Satuan': [np.nan, 819402.32, 81500, 98000],
                        'Jumlah': [np.nan] * 4
                    }
                    st.session_state.data_galian = data_galian
            elif jenis_galian == 'Galian Lumpur':
                if metode == 'Manual':
                    data_galian = {
                        'Uraian': ['Galian', 'Pekerja', 'Mandor'],
                        'Koefisien': [np.nan, 0.83, 0.083],
                        'Volume': [np.nan] + [volume_galian] * 2,
                        'Satuan': [np.nan, 'OH', 'OH'],
                        'Harga Satuan': [np.nan, 81500, 98000],
                        'Jumlah': [np.nan, np.nan, np.nan]
                    }
                    st.session_state.data_galian = data_galian
                elif metode == 'Semi Mekanis':
                    data_galian = {
                        'Uraian': ['Galian', 'Pertamina dex', 'Mesin Pompa Lumpur', 'Pekerja', 'Mandor'],
                        'Koefisien': [np.nan, 0.9, 0.0009, 0.24, 0.024],
                        'Volume': [np.nan] + [volume_galian] * 5,
                        'Satuan': [np.nan, 'liter', 'sewa-harian', 'OH', 'OH'],
                        'Harga Satuan': [np.nan, 13300, 339221000, 81500, 98000],
                        'Jumlah': [np.nan, np.nan, np.nan, np.nan, np.nan]
                    }
                    st.session_state.data_galian = data_galian
                elif metode == 'Mekanis':
                    data_galian = {
                        'Uraian': ['Galian', 'Dump Truck', 'Excavator', 'Pekerja', 'Mandor'],
                        'Koefisien': [np.nan, 0.07, 0.067, 0.226, 0.007],
                        'Volume': [np.nan] + [volume_galian] * 4,
                        'Satuan': [np.nan, 'jam', 'jam', 'OH', 'OH'],
                        'Harga Satuan': [np.nan, 339612.94, 819402.32, 81500, 98000],
                        'Jumlah': [np.nan, np.nan, np.nan, np.nan, np.nan]
                    }
                    st.session_state.data_galian = data_galian
            elif jenis_galian == 'Galian Cadas':
                if metode == 'Manual':
                    data_galian = {
                        'Uraian': ['Galian', 'Pekerja', 'Mandor'],
                        'Koefisien': [np.nan, 1.25, 0.125],
                        'Volume': [np.nan] + [volume_galian] * 2,
                        'Satuan': [np.nan, 'OH', 'OH'],
                        'Harga Satuan': [np.nan, 81500, 98000],
                        'Jumlah': [np.nan, np.nan, np.nan]
                    }
                    st.session_state.data_galian = data_galian
                elif metode == 'Semi Mekanis':
                    data_galian = {
                        'Uraian': ['Galian', 'Pertamina dex', 'Jack Hammer', 'Pekerja', 'Mandor'],
                        'Koefisien': [np.nan, 1.25, 0.125, 0.32, 0.032],
                        'Volume': [np.nan] + [volume_galian] * 4,
                        'Satuan': [np.nan, 'liter', 'sewa-harian', 'OH', 'OH'],
                        'Harga Satuan': [np.nan, 13300, 44700, 81500, 98000],
                        'Jumlah': [np.nan, np.nan, np.nan, np.nan, np.nan]
                    }
                    st.session_state.data_galian = data_galian
            elif jenis_galian == 'Galian Pasir':
                if metode == 'Manual':
                    data_galian = {
                        'Uraian': ['Galian', 'Pekerja', 'Mandor'],
                        'Koefisien': [np.nan, 0.66, 0.066],
                        'Volume': [np.nan] + [volume_galian] * 2,
                        'Satuan': [np.nan, 'OH', 'OH'],
                        'Harga Satuan': [np.nan, 81500, 98000],
                        'Jumlah': [np.nan, np.nan, np.nan]
                    }
                    st.session_state.data_galian = data_galian
                elif metode == 'Semi Mekanis':
                    data_galian = {
                        'Uraian': ['Galian', 'Pertamina dex', 'Mesin Pompa Lumpur', 'Pekerja', 'Mandor'],
                        'Koefisien': [np.nan, 1.1, 0.0002, 0.1, 0.01],
                        'Volume': [np.nan] + [volume_galian] * 4,
                        'Satuan': [np.nan, 'liter', 'sewa-harian', 'OH', 'OH'],
                        'Harga Satuan': [np.nan, 13300, 339221000, 81500, 98000],
                        'Jumlah': [np.nan, np.nan, np.nan, np.nan, np.nan]
                    }
                    st.session_state.data_galian = data_galian
        
            else:
                st.warning("Jenis galian tidak valid atau belum dipilih.")
                return


        # Definisi data untuk tabel urugan
            data_urugan = {
                    'Uraian': ['Urugan', 'Batu Belah', 'Batu Pecah Uk. 1-2 cm (mesin)', 'Pasir Lokal', 'Pekerja', 'Mandor'],
                    'Koefisien': [np.nan, 0.15, 0.09, 0.01, 1, 0.005],
                    'Volume': [np.nan] + [luas_urugan] * 5,
                    'Satuan': [np.nan, 'm3', 'm3', 'm3', 'OH', 'OH'],
                    'Harga Satuan': [np.nan, 198350, 325000, 233300, 81500, 98000],
                    'Jumlah': [np.nan, np.nan, np.nan, np.nan, np.nan, np.nan]
            }
            st.session_state.data_urugan = data_urugan

        # Definisi data untuk tabel pemadatan
            if pemadatan_urugan == 'Stamper':
                data_pemadatan = {
                    'Uraian': ['Pemadatan Tanah', 'Stamper', 'Pekerja', 'Tukang', 'Mandor'],
                    'Koefisien': [np.nan, 0.05, 0.062, 0.186, 0.0062],
                    'Volume': [np.nan] + [volume_pemadatan] * 4,
                    'Satuan': [np.nan, 'sewa-hari', 'OH', 'OH', 'OH'],
                    'Harga Satuan': [np.nan, 26050, 81500, 93000, 98000],
                    'Jumlah': [np.nan, np.nan, np.nan, np.nan, np.nan]
                }
                st.session_state.data_pemadatan = data_pemadatan
            elif pemadatan_urugan == 'Bulldozer':
                data_pemadatan = {
                    'Uraian': ['Pemadatan Tanah', 'Bulldozer', 'Water Tanker','Roller Vibrator','Pekerja', 'Mandor'],
                    'Koefisien': [np.nan, 0.02, 0.0078, 0.0178, 0.1422, 0.0142],
                    'Volume': [np.nan] + [volume_pemadatan] * 5,
                    'Satuan': [np.nan, 'jam', 'jam', 'jam', 'jam', 'jam'],
                    'Harga Satuan': [np.nan, 876843.85, 317296.73, 464580, 81500, 98000],
                    'Jumlah': [np.nan, np.nan, np.nan, np.nan, np.nan, np.nan]
                }
                st.session_state.data_pemadatan = data_pemadatan

    if st.session_state.get('show_estimasi_input', False):
        if st.button("Hitung Estimasi"):
            df_galian = pd.DataFrame(st.session_state.data_galian)
            df_galian['Jumlah'] = df_galian['Volume'] * df_galian['Koefisien'] * df_galian['Harga Satuan']
            rows_galian = list(dataframe_to_rows(df_galian, index=False, header=True))

            df_urugan = pd.DataFrame(st.session_state.data_urugan)
            df_urugan['Jumlah'] = df_galian['Volume'] * df_urugan['Koefisien'] * df_urugan['Harga Satuan']
            rows_urugan = list(dataframe_to_rows(df_urugan, index=False, header=False))

            df_pemadatan = pd.DataFrame(st.session_state.data_pemadatan)
            df_pemadatan['Jumlah'] = df_pemadatan['Volume'] * df_pemadatan['Koefisien'] * df_pemadatan['Harga Satuan']
            rows_pemadatan = list(dataframe_to_rows(df_pemadatan, index=False, header=False))

            # Membuat workbook dan worksheet baru
            wb = Workbook()
            ws = wb.active

            ws.title = "RAB Jalan Beton"
            cur_row = 1

            # Galian Data
            subtotal_galian = 0
            for r in rows_galian:
                ws.append(r)

                # calc subtotal
                if  isinstance(r[-1], (int, float)) and not np.isnan(r[-1]):
                    subtotal_galian += r[-1]

            ws.append(['subtotal', np.nan, np.nan, np.nan, np.nan, round(subtotal_galian, 0) ])
            if rows_galian:  # cek data galian tidak kosong
                ws.merge_cells(start_row=cur_row+1, start_column=1, end_row=cur_row+1, end_column=7)    
            cur_row = cur_row + len(rows_galian) + 1  # Update nilai cur_row setelah add data galian


            # Urugan Data
            subtotal_urugan = 0
            for r in rows_urugan:
                ws.append(r)

                # calc subtotal
                if  isinstance(r[-1], (int, float)) and not np.isnan(r[-1]):
                    subtotal_urugan += r[-1]

            ws.append(['subtotal', np.nan, np.nan, np.nan, np.nan, round(subtotal_urugan, 0) ])
            if rows_urugan:  # cek data urugan tidak kosong
                ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)       
            cur_row = cur_row + len(rows_urugan) + 1  # Update nilai cur_row setelah add data urugan


            # Pemadatan Data
            subtotal_pemadatan = 0
            for r in rows_pemadatan:
                ws.append(r)

                # calc subtotal
                if  isinstance(r[-1], (int, float)) and not np.isnan(r[-1]):
                    subtotal_pemadatan += r[-1]

            ws.append(['subtotal', np.nan, np.nan, np.nan, np.nan, round(subtotal_pemadatan, 0) ])
            if rows_pemadatan:  # cek data galian tidak kosong
                ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)          
            cur_row = cur_row + len(rows_pemadatan) + 1  # Update nilai cur_row setelah add data Pemadatan

            total = subtotal_galian + subtotal_urugan + subtotal_pemadatan
            ws.append(['TOTAL', np.nan, np.nan, np.nan, np.nan, round(total, 0) ])

            # Convert workbook to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            df = pd.read_excel(output)

            st.dataframe(df, hide_index=True)

            # Download the Excel file
            st.download_button(
                label="Download Excel",
                data=output,
                file_name="estimasi_rab_jalan_makadam.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        if st.button("Konfirmasi dan Selesai"):
            st.success("Estimasi RAB telah berhasil direkapitulasi.")
            st.balloons()

if __name__ == "__main__":
    st.title("Estimasi RAB")
    jalan_makadam_flow()
