import streamlit as st
import pandas as pd
from io import BytesIO
import numpy as np
import HSPK
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def jalan_paving_flow():

    st.subheader("Jalan Paving")

    # Menambahkan 3 kolom untuk gambar di bagian atas
    col1, col2 = st.columns(2)

    with col1:
        st.image("assets/paving 1.jpg", caption="1", use_column_width=True)

    with col2:
        st.image("assets/paving 2.jpg", caption="2", use_column_width=True)

    if st.button("Mulai Input"):
        st.session_state.jalan_paving = {}
        st.session_state.show_galian_input = True
        st.session_state.show_urugan_input = False
        st.session_state.show_paving_input = False
        st.session_state.show_estimasi_input = False

        st.session_state.rab_galian = None
        st.session_state.rab_pemadatan = None
        st.session_state.rab_pemasangan_paving = None
        st.session_state.rab_urugan = None

    # Bagian Galian
    if st.session_state.get('show_galian_input', False):
        st.write("### Pekerjaan Galian")
        jenis_galian = st.selectbox("Jenis Galian", ["Galian Batu", "Galian Tanah", "Galian Lumpur", "Galian Pasir", "Galian Cadas"])
        metode = st.radio("Metode Pekerjaan", ["Manual", "Mekanis", "Semi Mekanis"])
        panjang = st.number_input("Panjang Galian (m)", format="%.2f")
        lebar = st.number_input("Lebar Galian (m)", format="%.2f")
        kedalaman = st.number_input("Kedalaman Galian (m)", format="%.2f")

        if jenis_galian == "Galian Batu" and metode == "Manual":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_batu_manual = HSPK.GalianBatuManual(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_batu_manual.galian_batu_manual()
            st.session_state.rab_galian = rab_galian

        if jenis_galian == "Galian Tanah" and metode == "Manual":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_tanah_manual = HSPK.GalianTanahBiasaManual(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_tanah_manual.galian_tanah_biasa_manual()
            st.session_state.rab_galian = rab_galian
            
        if jenis_galian == "Galian Lumpur" and metode == "Manual":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_lumpur_manual = HSPK.GalianLumpurManual(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_lumpur_manual.galian_lumpur_manual()
            st.session_state.rab_galian = rab_galian

        if jenis_galian == "Galian Pasir" and metode == "Manual":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_pasir_manually = HSPK.GalianPasirManual(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_pasir_manually.galian_pasir_manual()
            st.session_state.rab_galian = rab_galian
            
        if jenis_galian == "Galian Cadas" and metode == "Manual":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_cadas_manual = HSPK.GalianTanahCadasManual(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_cadas_manual.galian_tanah_cadas_manual()
            st.session_state.rab_galian = rab_galian         

        if jenis_galian == "Galian Batu" and metode == "Mekanis":
            if kedalaman <= 1:
                volume_galian = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                volume_galian = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                volume_galian = 'galian_3m'
            else:
                volume_galian = 'penambahan_1m'

            galian_batu_mekanis = HSPK.GalianBatuDenganExcavator(panjang, lebar, kedalaman)
            rab_galian = galian_batu_mekanis.galian_biasa_dengan_excavator()
            st.session_state.rab_galian = rab_galian
            
        if jenis_galian == "Galian Tanah" and metode == "Mekanis":
            if kedalaman <= 1:
                volume_galian = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                volume_galian = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                volume_galian = 'galian_3m'
            else:
                volume_galian = 'penambahan_1m'

            galian_tanah_mekanis = HSPK.GalianTanahDenganExcavator(panjang, lebar, kedalaman)
            rab_galian = galian_tanah_mekanis.galian_tanah_dengan_excavator()
            st.session_state.rab_galian = rab_galian
            
        if jenis_galian == "Galian Lumpur" and metode == "Mekanis":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_lumpur_mekanis = HSPK.GalianLumpurMekanis(panjang, lebar, kedalaman)
            rab_galian = galian_lumpur_mekanis.galian_lumpur_mekanis()
            st.session_state.rab_galian = rab_galian
            
        if jenis_galian == "Galian Pasir" and metode == "Mekanis":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_pasir_manually = HSPK.GalianPasirManual(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_pasir_manually.galian_pasir_manual()
            st.session_state.rab_galian = rab_galian
            
        if jenis_galian == "Galian Cadas" and metode == "Mekanis":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_cadas_manual = HSPK.GalianTanahCadasManual(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_cadas_manual.galian_tanah_cadas_manual()
            st.session_state.rab_galian = rab_galian
            
        if jenis_galian == "Galian Batu" and metode == "Semi Mekanis":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_batu_semimekanis = HSPK.GalianBatuSemiMekanis(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_batu_semimekanis.galian_batu_semi_mekanis()
            st.session_state.rab_galian = rab_galian
            
        if jenis_galian == "Galian Tanah" and metode == "Semi Mekanis":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_tanah_semimekanis = HSPK.GalianTanahBiasaSemiMekanis(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_tanah_semimekanis.galian_tanah_biasa_semi_mekanis()
            st.session_state.rab_galian = rab_galian      

        if jenis_galian == "Galian Lumpur" and metode == "Semi Mekanis":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_lumpur_semimekanis = HSPK.GalianLumpurSemiMekanis(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_lumpur_semimekanis.galian_lumpur_semi_mekanis()
            st.session_state.rab_galian = rab_galian
            
        if jenis_galian == "Galian Pasir" and metode == "Semi Mekanis":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_pasir_semimekanis = HSPK.GalianPasirSemiMekanis(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_pasir_semimekanis.galian_pasir_semi_mekanis()
            st.session_state.rab_galian = rab_galian
            
        if jenis_galian == "Galian Cadas" and metode == "Semi Mekanis":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_cadas_semimekanis = HSPK.GalianTanahCadasSemiMekanis(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_cadas_semimekanis.galian_tanah_cadas_semi_mekanis()
            st.session_state.rab_galian = rab_galian     

        if st.button("Konfirmasi Galian", key="konfirmasi_galian"):
            if panjang == 0 or lebar == 0:
                st.error("Mohon masukkan panjang, lebar, dan kedalaman yang valid sebelum melanjutkan.")
            else:
                st.session_state.jalan_paving.update({
                    'jenis_galian': jenis_galian,
                    'metode': metode,
                    'panjang': panjang,
                    'lebar': lebar,
                    'kedalaman': kedalaman,
                })
                st.success("Input Galian disimpan! Lanjutkan ke Urugan.")
                st.session_state.show_urugan_input = True
                st.session_state.show_galian_input = False

    # Bagian Urugan
    if st.session_state.get('show_urugan_input', False):
        st.write("### Pekerjaan Urugan")
        panjang_urugan = st.number_input("Panjang Urugan (m)", format="%.2f")
        lebar_urugan = st.number_input("Lebar Urugan (m)", format="%.2f")
        kedalaman_urugan = st.number_input("Kedalaman Urugan (m)", format="%.2f")
        pemadatan_urugan = st.selectbox("Jenis Pemadatan", ["Stamper", "Bulldozer"])

        urugan_sirtu_padat = HSPK.Urugan(panjang_urugan, lebar_urugan, kedalaman_urugan, 'sirtu_padat')
        rab_urugan = urugan_sirtu_padat.urugan()
        st.session_state.rab_urugan = rab_urugan
        
        if pemadatan_urugan == "Stamper":
            pemadatan_stamper = HSPK.PemadatanTanahDenganStamper(panjang_urugan, lebar_urugan, kedalaman_urugan)
            rab_pemadatan = pemadatan_stamper.pemadatan_tanah_dengan_stamper()
            st.session_state.rab_pemadatan = rab_pemadatan
            
        else:
            pemadatan_bulldozer = HSPK.PemadatanTanahDenganBulldozer(panjang_urugan, lebar_urugan, kedalaman_urugan)
            rab_pemadatan = pemadatan_bulldozer.pemadatan_dengan_stamper()
            st.session_state.rab_pemadatan = rab_pemadatan
            
        if st.button("Konfirmasi Urugan", key="konfirmasi_urugan"):
            if panjang_urugan == 0 or lebar_urugan == 0:
                st.error("Mohon masukkan panjang dan lebar kedalaman yang valid sebelum melanjutkan.")
            else:
                st.session_state.jalan_paving.update({
                    'panjang_urugan': panjang_urugan,
                    'lebar_urugan': lebar_urugan,
                    'kedalaman_urugan': lebar_urugan,
                    'pemadatan_urugan': pemadatan_urugan,
                })
                st.success("Input Urugan disimpan! Lanjutkan ke Pemasangan Paving.")
                st.session_state.show_paving_input = True
                st.session_state.show_urugan_input = False

    # Pemasangan Paving: This step is mandatory after both Galian and Urugan
    if st.session_state.get('show_paving_input', False):
        st.write("### Pemasangan Paving")
        panjang_paving = st.number_input("Panjang Daerah Paving (m)", format="%.2f", )
        lebar_paving = st.number_input("Lebar Daerah Paving (m)", format="%.2f")
        jenis_paving = st.selectbox("Jenis Paving", ["Paving Berwarna", "Paving Natural"])
        ketebalan_paving = st.selectbox("Ketebalan Paving", ["6 cm", "8 cm"])

        if jenis_paving == "Paving Berwarna" and ketebalan_paving == "6 cm":
            tipe_koefisien = 'paving_berwarna_6cm' 
        elif jenis_paving == "Paving Berwarna" and ketebalan_paving == "8 cm":
            tipe_koefisien = 'paving_berwarna_8cm'
        elif jenis_paving == "Paving Natural" and ketebalan_paving == "6 cm":
            tipe_koefisien = 'paving_natural_6cm'
        elif jenis_paving == "Paving Natural" and ketebalan_paving == "8 cm":
            tipe_koefisien = 'paving_natural_8cm'

        pemasangan_paving = HSPK.PemasanganPaving(panjang_paving, lebar_paving, tipe_koefisien)
        rab_pemasangan_paving = pemasangan_paving.hasil_perhitungan()
        st.session_state.rab_pemasangan_paving = rab_pemasangan_paving

        if st.button("Konfirmasi Pemasangan Paving", key="Konfirmasi_Pemasangan_Paving"):
            st.session_state.jalan_paving.update({
                'panjang_paving': panjang_paving,
                'lebar_paving': lebar_paving,
                'Jenis_paving': jenis_paving,
                'ketebalan_paving': ketebalan_paving,
            })
            st.success("Input Paving disimpan! Tekan tombol Submit untuk menyimpan semua data.")
            st.session_state.show_estimasi_input = True
            st.session_state.show_paving_input = False

    # Bagian Perhitungan Estimasi RAB
    if st.session_state.get('show_estimasi_input', False):

        st.write("### Perhitungan Estimasi RAB Pembuatan Jalan Paving")

        if st.button("Hitung Estimasi"):
            df_galian = pd.DataFrame(st.session_state.rab_galian['data'])
            rows_galian = list(dataframe_to_rows(df_galian, index=False, header=True))

            df_urugan = pd.DataFrame(st.session_state.rab_urugan['data'])
            rows_urugan = list(dataframe_to_rows(df_urugan, index=False, header=False))

            df_pemadatan = pd.DataFrame(st.session_state.rab_pemadatan['data'])
            rows_pemadatan = list(dataframe_to_rows(df_pemadatan, index=False, header=False))

            df_pemasangan_paving = pd.DataFrame(st.session_state.rab_pemasangan_paving['data'])
            rows_pemasangan_paving = list(dataframe_to_rows(df_pemasangan_paving, index=False, header=False))

            # Membuat workbook dan worksheet baru
            wb = Workbook()
            ws = wb.active

            ws.title = "RAB Jalan Paving"
            cur_row = 1

            # Galian Data
            subtotal_galian = 0
            for r in rows_galian:
                ws.append(r)

                # calc subtotal
                if  isinstance(r[-1], (int, float)) and not np.isnan(r[-1]):
                    subtotal_galian += r[-1]

            ws.append(['subtotal', np.nan, np.nan, np.nan, np.nan, round(subtotal_galian, 0) ])
            if rows_galian:  # cek data galian tidak kosong
                ws.merge_cells(start_row=cur_row+1, start_column=1, end_row=cur_row+1, end_column=7)    
            cur_row = cur_row + len(rows_galian) + 1  # Update nilai cur_row setelah add data galian


            # Urugan Data
            subtotal_urugan = 0
            for r in rows_urugan:
                ws.append(r)

                # calc subtotal
                if  isinstance(r[-1], (int, float)) and not np.isnan(r[-1]):
                    subtotal_urugan += r[-1]

            ws.append(['subtotal', np.nan, np.nan, np.nan, np.nan, round(subtotal_urugan, 0) ])
            if rows_urugan:  # cek data urugan tidak kosong
                ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)       
            cur_row = cur_row + len(rows_urugan) + 1  # Update nilai cur_row setelah add data urugan


            # Pemadatan Data
            subtotal_pemadatan = 0
            for r in rows_pemadatan:
                ws.append(r)

                # calc subtotal
                if  isinstance(r[-1], (int, float)) and not np.isnan(r[-1]):
                    subtotal_pemadatan += r[-1]

            ws.append(['subtotal', np.nan, np.nan, np.nan, np.nan, round(subtotal_pemadatan, 0) ])
            if rows_pemadatan:  # cek data galian tidak kosong
                ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)          
            cur_row = cur_row + len(rows_pemadatan) + 1  # Update nilai cur_row setelah add data Pemadatan


            # Pemasangan Paving Data
            subtotal_pemasangan_paving = 0
            for r in rows_pemasangan_paving:
                ws.append(r)

                # calc subtotat
                if  isinstance(r[-1], (int, float)) and not np.isnan(r[-1]):
                    subtotal_pemasangan_paving += r[-1]

            ws.append(['subtotal', np.nan, np.nan, np.nan, np.nan, round(subtotal_pemasangan_paving, 0) ])
            if rows_pemasangan_paving:  # cek data galian tidak kosong
                ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)   
            cur_row = cur_row + len(rows_pemasangan_paving) + 1  # Update nilai cur_row setelah add data pemasangan paving

            total = subtotal_galian + subtotal_urugan + subtotal_pemadatan + subtotal_pemasangan_paving
            ws.append(['TOTAL', np.nan, np.nan, np.nan, np.nan, round(total, 0) ])

            # Convert workbook to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            df = pd.read_excel(output)

            st.dataframe(df, hide_index=True)

            # Download the Excel file
            st.download_button(
                label="Download Excel",
                data=output,
                file_name="estimasi_rab_jalan_paving.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        if st.button("Konfirmasi dan Selesai"):
            st.success("RAB telah berhasil direkapitulasi dan selesai.")
            st.balloons()


if __name__ == "__main__":
    st.title("Estimasi RAB")
    jalan_paving_flow()
