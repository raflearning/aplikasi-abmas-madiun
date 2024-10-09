import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import HSPK
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def saluran_drainase_flow():
    st.subheader("Saluran Drainase")

    # Menambahkan 3 kolom untuk gambar di bagian atas
    col1, col2 = st.columns(2)

    with col1:
        st.image("assets/drainase 1.jpg", caption="1", use_column_width=True)

    with col2:
        st.image("assets/drainase 2.jpg", caption="2", use_column_width=True)

    if st.button("Mulai Input"):
        st.session_state.jalan_beton = {}
        st.session_state.show_galian_input = True
        st.session_state.show_beton_a_input = False
        st.session_state.show_beton_b_input = False
        st.session_state.show_urugan_input = False
        st.session_state.show_pemadatan_input = False
        st.session_state.show_estimasi_input = False

        st.session_state.rab_galian = None
        st.session_state.rab_beton_a = None
        st.session_state.rab_beton_b = None
        st.session_state.rab_perancah_a = None
        st.session_state.rab_perancah_a = None
        st.session_state.rab_urugan = None
        st.session_state.rab_pemadatan = None
        

    # Bagian Galian
    if st.session_state.get('show_galian_input', False):
        st.write("### Pekerjaan Galian")
        jenis_galian = st.selectbox("Jenis Galian", ["Galian Batu", "Galian Tanah", "Galian Lumpur", "Galian Pasir", "Galian Cadas"])
        metode = st.radio("Metode Pekerjaan", ["Manual", "Mekanis", "Semi Mekanis"])

        panjang = st.number_input("Panjang Galian (m)", format="%.2f")
        lebar = st.number_input("Lebar Galian (m)", format="%.2f")
        kedalaman = st.number_input("Kedalaman Galian (m)", format="%.2f")

        if jenis_galian == "Galian Batu" and metode == "Manual":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_batu_manual = HSPK.GalianBatuManual(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_batu_manual.galian_batu_manual()
            st.session_state.rab_galian = rab_galian

        if jenis_galian == "Galian Tanah" and metode == "Manual":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_tanah_manual = HSPK.GalianTanahBiasaManual(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_tanah_manual.galian_tanah_biasa_manual()
            st.session_state.rab_galian = rab_galian

        if jenis_galian == "Galian Lumpur" and metode == "Manual":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_lumpur_manual = HSPK.GalianLumpurManual(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_lumpur_manual.galian_lumpur_manual()
            st.session_state.rab_galian = rab_galian

        if jenis_galian == "Galian Pasir" and metode == "Manual":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_pasir_manually = HSPK.GalianPasirManual(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_pasir_manually.galian_pasir_manual()
            st.session_state.rab_galian = rab_galian

        if jenis_galian == "Galian Cadas" and metode == "Manual":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_cadas_manual = HSPK.GalianTanahCadasManual(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_cadas_manual.galian_tanah_cadas_manual()
            st.session_state.rab_galian = rab_galian

        if jenis_galian == "Galian Batu" and metode == "Mekanis":
            if kedalaman <= 1:
                volume_galian = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                volume_galian = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                volume_galian = 'galian_3m'
            else:
                volume_galian = 'penambahan_1m'

            galian_batu_mekanis = HSPK.GalianBatuDenganExcavator(panjang, lebar, kedalaman)
            rab_galian = galian_batu_mekanis.galian_biasa_dengan_excavator()
            st.session_state.rab_galian = rab_galian

        if jenis_galian == "Galian Tanah" and metode == "Mekanis":
            if kedalaman <= 1:
                volume_galian = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                volume_galian = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                volume_galian = 'galian_3m'
            else:
                volume_galian = 'penambahan_1m'

            galian_tanah_mekanis = HSPK.GalianTanahDenganExcavator(panjang, lebar, kedalaman)
            rab_galian = galian_tanah_mekanis.galian_tanah_dengan_excavator()
            st.session_state.rab_galian = rab_galian

        if jenis_galian == "Galian Lumpur" and metode == "Mekanis":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_lumpur_mekanis = HSPK.GalianLumpurMekanis(panjang, lebar, kedalaman)
            rab_galian = galian_lumpur_mekanis.galian_lumpur_mekanis()
            st.session_state.rab_galian = rab_galian

        if jenis_galian == "Galian Pasir" and metode == "Mekanis":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_pasir_manually = HSPK.GalianPasirManual(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_pasir_manually.galian_pasir_manual()
            st.session_state.rab_galian = rab_galian

        if jenis_galian == "Galian Cadas" and metode == "Mekanis":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_cadas_manual = HSPK.GalianTanahCadasManual(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_cadas_manual.galian_tanah_cadas_manual()
            st.session_state.rab_galian = rab_galian

        if jenis_galian == "Galian Batu" and metode == "Semi Mekanis":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_batu_semimekanis = HSPK.GalianBatuSemiMekanis(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_batu_semimekanis.galian_batu_semi_mekanis()
            st.session_state.rab_galian = rab_galian

        if jenis_galian == "Galian Tanah" and metode == "Semi Mekanis":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_tanah_semimekanis = HSPK.GalianTanahBiasaSemiMekanis(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_tanah_semimekanis.galian_tanah_biasa_semi_mekanis()
            st.session_state.rab_galian = rab_galian

        if jenis_galian == "Galian Lumpur" and metode == "Semi Mekanis":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_lumpur_semimekanis = HSPK.GalianLumpurSemiMekanis(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_lumpur_semimekanis.galian_lumpur_semi_mekanis()
            st.session_state.rab_galian = rab_galian

        if jenis_galian == "Galian Pasir" and metode == "Semi Mekanis":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_pasir_semimekanis = HSPK.GalianPasirSemiMekanis(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_pasir_semimekanis.galian_pasir_semi_mekanis()
            st.session_state.rab_galian = rab_galian

        if jenis_galian == "Galian Cadas" and metode == "Semi Mekanis":
            if kedalaman <= 1:
                tipe_koefisien = 'galian_1m'
            elif kedalaman >= 1 and kedalaman < 2:
                tipe_koefisien = 'galian_2m'
            elif kedalaman >= 2 and kedalaman < 3:
                tipe_koefisien = 'galian_3m'
            else:
                tipe_koefisien = 'penambahan_1m'

            galian_cadas_semimekanis = HSPK.GalianTanahCadasSemiMekanis(panjang, lebar, kedalaman, tipe_koefisien)
            rab_galian = galian_cadas_semimekanis.galian_tanah_cadas_semi_mekanis()
            st.session_state.rab_galian = rab_galian

        if st.button("Konfirmasi Galian", key="konfirmasi_galian"):
            if panjang == 0 or lebar == 0:
                st.error("Mohon masukkan panjang, lebar, dan kedalaman yang valid sebelum melanjutkan.")
            else:
                st.session_state.jalan_beton.update({
                    'jenis_galian': jenis_galian,
                    'metode': metode,
                    'panjang': panjang,
                    'lebar': lebar,
                    'kedalaman': kedalaman,
                })
                st.success("Input Galian disimpan! Lanjutkan ke Beton Sisi A.")
                st.session_state.show_beton_a_input = True
                st.session_state.show_galian_input = False        

    #Bagian Beton A
    if st.session_state.get('show_beton_a_input', False):
        st.write("### Pembuatan Beton Darinase Bagian (A)")
        mutu_beton = st.selectbox("Mutu Beton", ["K125", "K150", "K175", "K225", "K250", "K300"], key = 'mutua')
        metode_perancah = st.selectbox("Bekisting & Perancah", ["Kayu Jawa", "Dolken", "Bambu"], key = 'perancaha')
        panjang_beton = st.number_input("Panjang Beton (m)", format="%.2f", key = 'panjanga')
        lebar_beton = st.number_input("Lebar Beton (m)", format="%.2f", key = 'lebara')
        ketebalan_beton = st.number_input("Ketebalan Beton (m)", format="%.2f", key = 'tebala')
        ketebalan_bekisting = st.selectbox("Jenis Bekisting", ['multipleks 12mm', 'multipleks 18mm'], key = 'bekistinga')

        if mutu_beton == "K125":
            tipe_koefisien = 'beton_k125'
        elif mutu_beton == "K150":
            tipe_koefisien = 'beton_k150'
        elif mutu_beton == "K175":
            tipe_koefisien = 'beton_k175'
        elif mutu_beton == "K225":
            tipe_koefisien = 'beton_k225'
        elif mutu_beton == "K250":
            tipe_koefisien = 'beton_k250'
        else:
            tipe_koefisien = 'beton_k300'

        pemilihan_mutu = HSPK.BetonReadyMix(panjang_beton, lebar_beton, ketebalan_beton, tipe_koefisien)
        rab_beton_a = pemilihan_mutu.beton_readymix()
        st.session_state.rab_beton_a = rab_beton_a

        if ketebalan_bekisting == 'multipleks 12mm':
            tipe_koefisien = 'multipleks_12mm'
        else:
            tipe_koefisien = 'multipleks_18mm'

        bekisting_lantai = HSPK.BekistingLantai(panjang_beton, lebar_beton, tipe_koefisien)
        rab_bekisting_a = bekisting_lantai.bekisting_lantai()
        st.session_state.rab_bekisting_a = rab_bekisting_a
        
        if metode_perancah == "Kayu Jawa":
            perancah_kayu = HSPK.PerancahLantaiKayu(panjang_beton, lebar_beton)
            rab_perancah_a = perancah_kayu.ahsp_perancah_lantai()
            st.session_state.rab_perancah_a = rab_perancah_a

        elif metode_perancah == "Dolken":

            perancah_dolken = HSPK.PerancahLantaiDolken(panjang_beton, lebar_beton)
            rab_perancahA = perancah_dolken.ahsp_perancah_lantai()
            st.session_state.rab_perancahA = rab_perancahA

        else:
            perancah_bambu = HSPK.PerancahLantaiBambu(panjang_beton, lebar_beton)
            rab_perancah_a = perancah_bambu.ahsp_perancah_lantai()
            st.session_state.rab_perancah_a = rab_perancah_a

        if st.button("Konfirmasi Beton Sisi A", key="Konfirmasi_Pembuatan_Beton_a"):
            st.session_state.jalan_beton.update({
                'panjang_beton': panjang_beton,
                'lebar_beton': lebar_beton,
                'ketebalan_beton': ketebalan_beton,
                'mutu_beton': mutu_beton,
                'ketebalan_bekisting' : ketebalan_bekisting,
                'metode_perancah': metode_perancah})
            st.success("Input Beton Sisi A disimpan! Lanjut ke Beton Sisi B")
            st.session_state.show_beton_b_input = True
            st.session_state.show_beton_a_input = False

    #Bagian Beton B
    if st.session_state.get('show_beton_b_input', False):
        st.write("### Pembuatan Beton Darinase Bagian (B)")
        mutu_beton = st.selectbox("Mutu Beton", ["K125", "K150", "K175", "K225", "K250", "K300"], key = 'mutub')
        metode_perancah = st.selectbox("Bekisting & Perancah", ["Kayu Jawa", "Dolken", "Bambu"], key = 'perancahb')
        panjang_beton = st.number_input("Panjang Beton (m)", format="%.2f", key = 'panjangb')
        lebar_beton = st.number_input("Lebar Beton (m)", format="%.2f", key = 'lebarb')
        ketebalan_beton = st.number_input("Ketebalan Beton (m)", format="%.2f", key = 'tebab')
        ketebalan_bekisting = st.selectbox("Jenis Bekisting", ['multipleks 12mm', 'multipleks 18mm'], key = 'bekistingb')

        if mutu_beton == "K125":
            tipe_koefisien = 'beton_k125'
        elif mutu_beton == "K150":
            tipe_koefisien = 'beton_k150'
        elif mutu_beton == "K175":
            tipe_koefisien = 'beton_k175'
        elif mutu_beton == "K225":
            tipe_koefisien = 'beton_k225'
        elif mutu_beton == "K250":
            tipe_koefisien = 'beton_k250'
        else:
            tipe_koefisien = 'beton_k300'

        pemilihan_mutu = HSPK.BetonReadyMix(panjang_beton, lebar_beton, ketebalan_beton, tipe_koefisien)
        rab_beton_b = pemilihan_mutu.beton_readymix()
        st.session_state.rab_beton_b = rab_beton_b

        if ketebalan_bekisting == 'multipleks 12mm':
            tipe_koefisien = 'multipleks_12mm'
        else: 
            tipe_koefisien = 'multipleks_18mm'

        bekisting_lantai = HSPK.BekistingLantai(panjang_beton, lebar_beton, tipe_koefisien)
        rab_bekisting_b = bekisting_lantai.bekisting_lantai()
        st.session_state.rab_bekisting_b = rab_bekisting_b
        
        if metode_perancah == "Kayu Jawa":
            perancah_kayu = HSPK.PerancahLantaiKayu(panjang_beton, lebar_beton)
            rab_perancah_b = perancah_kayu.ahsp_perancah_lantai()
            st.session_state.rab_perancah_b = rab_perancah_b

        elif metode_perancah == "Dolken":

            perancah_dolken = HSPK.PerancahLantaiDolken(panjang_beton, lebar_beton)
            rab_perancah_b = perancah_dolken.ahsp_perancah_lantai()
            st.session_state.rab_perancah_b = rab_perancah_b

        else:
            perancah_bambu = HSPK.PerancahLantaiBambu(panjang_beton, lebar_beton)
            rab_perancah_b = perancah_bambu.ahsp_perancah_lantai()
            st.session_state.rab_perancah_b = rab_perancah_b

        if st.button("Konfirmasi Beton Sisi B", key="Konfirmasi_Pembuatan_Beton_b"):
            st.session_state.jalan_beton.update({
                'panjang_beton': panjang_beton,
                'lebar_beton': lebar_beton,
                'ketebalan_beton': ketebalan_beton,
                'mutu_beton': mutu_beton,
                'ketebalan_bekisting' : ketebalan_bekisting,
                'metode_perancah': metode_perancah})
            st.success("Input Pembuatan Beton Sisi B disimpan! Lanjut ke Urugan")
            st.session_state.show_urugan_input = True
            st.session_state.show_beton_b_input = False

    # Bagian Urugan
    if st.session_state.get('show_urugan_input', False):
        st.write("### Pekerjaan Urugan")
        panjang_urugan = st.number_input("Panjang Urugan (m)", format="%.2f")
        lebar_urugan = st.number_input("Lebar Urugan (m)", format="%.2f")
        kedalaman_urugan = st.number_input("Kedalaman Urugan (m)", format="%.2f")
        pemadatan_urugan = st.selectbox("Jenis Pemadatan", ["Stamper", "Bulldozer"])

        urugan_sirtu_padat = HSPK.Urugan(panjang_urugan, lebar_urugan, kedalaman_urugan, 'sirtu_padat')
        rab_urugan = urugan_sirtu_padat.urugan()
        st.session_state.rab_urugan = rab_urugan

        if pemadatan_urugan == "Stamper":
            pemadatan_stamper = HSPK.PemadatanTanahDenganStamper(panjang_urugan, lebar_urugan, kedalaman_urugan)
            rab_pemadatan = pemadatan_stamper.pemadatan_tanah_dengan_stamper()
            st.session_state.rab_pemadatan = rab_pemadatan

        else:
            pemadatan_bulldozer = HSPK.PemadatanTanahDenganBulldozer(panjang_urugan, lebar_urugan, kedalaman_urugan)
            rab_pemadatan = pemadatan_bulldozer.pemadatan_dengan_stamper()
            st.session_state.rab_pemadatan = rab_pemadatan

        if st.button("Konfirmasi Urugan", key="konfirmasi_urugan"):
            if panjang_urugan == 0 or lebar_urugan == 0:
                st.error("Mohon masukkan panjang dan lebar kedalaman yang valid sebelum melanjutkan.")
            else:
                st.session_state.jalan_beton.update({
                    'panjang_urugan': panjang_urugan,
                    'lebar_urugan': lebar_urugan,
                    'kedalaman_urugan': lebar_urugan,
                    'pemadatan_urugan': pemadatan_urugan,
                })
                st.success("Input Urugan disimpan! Dapatkan Estimasinya.")
                st.session_state.show_estimasi_input = True
                st.session_state.show_urugan_input = False

    # Bagian Perhitungan Estimasi RAB
    if st.session_state.get('show_estimasi_input', False):
        st.write("### Perhitungan Estimasi RAB Pembuatan Saluran")

        if st.button("Hitung Estimasi"):
            df_galian = pd.DataFrame(st.session_state.rab_galian['data'])
            rows_galian = list(dataframe_to_rows(df_galian, index=False, header=True))

            df_beton_a = pd.DataFrame(st.session_state.rab_beton_a['data'])
            rows_beton_a = list(dataframe_to_rows(df_beton_a, index=False, header=True))
 
            df_beton_b = pd.DataFrame(st.session_state.rab_beton_b['data'])
            rows_beton_b = list(dataframe_to_rows(df_beton_b, index=False, header=True))

            df_perancah_a = pd.DataFrame(st.session_state.rab_perancah_a['data'])
            rows_perancah_a = list(dataframe_to_rows(df_perancah_a, index=False, header=True))
      
            df_perancah_b = pd.DataFrame(st.session_state.rab_perancah_b['data'])
            rows_perancah_b = list(dataframe_to_rows(df_perancah_b, index=False, header=True))

            df_urugan = pd.DataFrame(st.session_state.rab_urugan['data'])
            rows_urugan = list(dataframe_to_rows(df_urugan, index=False, header=True))

            df_pemadatan = pd.DataFrame(st.session_state.rab_pemadatan['data'])
            rows_pemadatan = list(dataframe_to_rows(df_pemadatan, index=False, header=True))
            
            # Membuat workbook dan worksheet baru
            wb = Workbook()
            ws = wb.active
            ws.title = "RAB Drainase"
            cur_row = 1
            
            # Galian Data
            subtotal_galian = 0
            for r in rows_galian:
                ws.append(r)

                # Cacl subtotal
                if isinstance(r[-1], (int, float)) and not np.isnan(r[-1]):
                    subtotal_galian += r[-1]
            
            ws.append(['subtotal', np.nan, np.nan, np.nan, np.nan, round(subtotal_galian, 0) ])
            if rows_galian: # Cek data galian tidak kosong
                ws.merge_cells(start_row=cur_row+1, start_column=1, end_row=cur_row+1, end_column=7)    
            cur_row = cur_row + len(rows_galian) + 1  # Update nilai cur_row setelah add data galian
            
            #  Beton A Data
            subtotal_beton_a = 0
            for r in rows_beton_a:
                ws.append(r)

                # Cacl subtotal
                if isinstance(r[-1], (int, float)) and not np.isnan(r[-1]):
                    subtotal_beton_a += r[-1]
            
            ws.append(['subtotal', np.nan, np.nan, np.nan, np.nan, round(subtotal_beton_a, 0) ])
            if rows_beton_a: # Cek data beton a tidak kosong
                ws.merge_cells(start_row=cur_row+1, start_column=1, end_row=cur_row+1, end_column=7)    
            cur_row = cur_row + len(rows_beton_a) + 1  # Update nilai cur_row setelah add data beton a

            # Beton B Data
            subtotal_beton_b = 0
            for r in rows_beton_b:
                ws.append(r)

                # Cacl subtotal
                if isinstance(r[-1], (int, float)) and not np.isnan(r[-1]):
                    subtotal_beton_b += r[-1]
            
            ws.append(['subtotal', np.nan, np.nan, np.nan, np.nan, round(subtotal_beton_b, 0) ])
            if rows_beton_b: # Cek data beton b tidak kosong
                ws.merge_cells(start_row=cur_row+1, start_column=1, end_row=cur_row+1, end_column=7)    
            cur_row = cur_row + len(rows_beton_b) + 1  # Update nilai cur_row setelah add data beton b

            #  Perancah A Data
            subtotal_perancah_a = 0
            for r in rows_perancah_a:
                ws.append(r)

                # Cacl subtotal
                if isinstance(r[-1], (int, float)) and not np.isnan(r[-1]):
                    subtotal_perancah_a += r[-1]
            
            ws.append(['subtotal', np.nan, np.nan, np.nan, np.nan, round(subtotal_perancah_a, 0) ])
            if rows_perancah_a: # Cek data perancah a tidak kosong
                ws.merge_cells(start_row=cur_row+1, start_column=1, end_row=cur_row+1, end_column=7)    
            cur_row = cur_row + len(rows_perancah_a) + 1  # Update nilai cur_row setelah add data perancah a

            # Perancah B Data
            subtotal_perancah_b = 0
            for r in rows_perancah_b:
                ws.append(r)

                # Cacl subtotal
                if isinstance(r[-1], (int, float)) and not np.isnan(r[-1]):
                    subtotal_perancah_b += r[-1]
            
            ws.append(['subtotal', np.nan, np.nan, np.nan, np.nan, round(subtotal_perancah_b, 0) ])
            if rows_perancah_b: # Cek data perancah b tidak kosong
                ws.merge_cells(start_row=cur_row+1, start_column=1, end_row=cur_row+1, end_column=7)    
            cur_row = cur_row + len(rows_perancah_b) + 1  # Update nilai cur_row setelah add data perancah b

            # Urugan Data
            subtotal_urugan = 0
            for r in rows_urugan:
                ws.append(r)

                # Cacl subtotal
                if isinstance(r[-1], (int, float)) and not np.isnan(r[-1]):
                    subtotal_urugan += r[-1]
            
            ws.append(['subtotal', np.nan, np.nan, np.nan, np.nan, round(subtotal_urugan, 0)])
            if rows_urugan: # Cek data urugan tidak kosong
                ws.merge_cells(start_row=cur_row+1, start_column=1, end_row=cur_row+1, end_column=7)    
            cur_row = cur_row + len(rows_urugan) + 1  # Update nilai cur_row setelah add data urugan

            # Pemadatan Data
            subtotal_pemadatan = 0
            for r in rows_pemadatan:
                ws.append(r)

                # Cacl subtotal
                if isinstance(r[-1], (int, float)) and not np.isnan(r[-1]):
                    subtotal_pemadatan += r[-1]
            
            ws.append(['subtotal', np.nan, np.nan, np.nan, np.nan, subtotal_pemadatan])
            if rows_pemadatan: # Cek data pemadatan tidak kosong
                ws.merge_cells(start_row=cur_row+1, start_column=1, end_row=cur_row+1, end_column=7)    
            cur_row = cur_row + len(rows_pemadatan) + 1  # Update nilai cur_row setelah add data pemadatan

            # Total Data
            total = subtotal_galian + subtotal_beton_a + subtotal_beton_b + subtotal_perancah_a + subtotal_perancah_b + subtotal_urugan + subtotal_pemadatan
            ws.append(['TOTAL', np.nan, np.nan, np.nan, np.nan, round(total, 0) ])
            
            # Convert workbook to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            df = pd.read_excel(output)

            st.dataframe(df, hide_index=True)

            # Download the Excel file
            st.download_button(
                label="Download Excel",
                data=output,
                file_name="estimasi_rab_drainase.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        if st.button("Konfirmasi dan Selesai"):
            st.success("RAB telah berhasil direkapitulasi dan selesai.")
            st.balloons()

if __name__ == "__main__":
    st.title("Estimasi RAB")
    saluran_drainase_flow()

